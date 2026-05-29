import json
import logging
import os
import re
import shutil
from datetime import datetime
from typing import Any, Dict, List, Optional

logger = logging.getLogger("evidence_extractor")
logging.basicConfig(level=logging.INFO)

try:
    from PIL import Image, ExifTags
    PIL_AVAILABLE = True
except Exception:
    PIL_AVAILABLE = False
    Image = None
    ExifTags = None

try:
    from PyPDF2 import PdfReader
    PDF_AVAILABLE = True
except Exception:
    PDF_AVAILABLE = False
    PdfReader = None

try:
    from docx import Document as DocxDocument
    DOCX_AVAILABLE = True
except Exception:
    DOCX_AVAILABLE = False
    DocxDocument = None

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False
    load_workbook = None

try:
    import xlrd
    XLRD_AVAILABLE = True
except Exception:
    XLRD_AVAILABLE = False
    xlrd = None


class EvidenceExtractor:
    """
    Stores claim evidence on disk and extracts lightweight metadata for analysis.
    """

    ROOT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    BASE_DIR = os.path.join(ROOT_DIR, "data", "uploads", "claims")

    MIN_PHOTOS = 1
    MAX_PHOTOS = 7
    MAX_PHOTO_BYTES = 10 * 1024 * 1024
    MAX_PDF_BYTES = 20 * 1024 * 1024

    ALLOWED_IMAGE_TYPES = {"image/jpeg", "image/png", "image/webp"}
    ALLOWED_IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp"}
    ALLOWED_PDF_TYPES = {
        "application/pdf",
        "application/msword",
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/vnd.ms-excel",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }
    ALLOWED_PDF_EXTS = {".pdf", ".doc", ".docx", ".xls", ".xlsx"}

    @classmethod
    def ensure_claim_dirs(cls, claim_id: str) -> Dict[str, str]:
        base = os.path.join(cls.BASE_DIR, claim_id)
        images_dir = os.path.join(base, "images")
        docs_dir = os.path.join(base, "docs")
        os.makedirs(images_dir, exist_ok=True)
        os.makedirs(docs_dir, exist_ok=True)
        return {"base": base, "images": images_dir, "docs": docs_dir}

    @classmethod
    def purge_claim_dir(cls, claim_id: str) -> None:
        base = os.path.join(cls.BASE_DIR, claim_id)
        if os.path.exists(base):
            shutil.rmtree(base, ignore_errors=True)

    @classmethod
    def validate_evidence(cls, photos: List[Any], pdf_file: Optional[Any]) -> None:
        if len(photos) < cls.MIN_PHOTOS or len(photos) > cls.MAX_PHOTOS:
            raise ValueError(
                f"Se requieren entre {cls.MIN_PHOTOS} y {cls.MAX_PHOTOS} fotos. Recibidas: {len(photos)}."
            )

        for photo in photos:
            if not cls._is_allowed_image(photo):
                raise ValueError("Formato de imagen no soportado. Use JPG, PNG o WEBP.")

            size_bytes = cls._get_upload_size(photo)
            if size_bytes > cls.MAX_PHOTO_BYTES:
                raise ValueError("Una foto supera el tamanio maximo permitido (10 MB).")

        if pdf_file is None:
            return  # PDF is optional — skip document validation if not provided

        if not cls._is_allowed_pdf(pdf_file):
            raise ValueError("Formato de documento no soportado. Use PDF, Word o Excel.")

        pdf_size = cls._get_upload_size(pdf_file)
        if pdf_size > cls.MAX_PDF_BYTES:
            raise ValueError("El documento supera el tamanio maximo permitido (20 MB).")

    @classmethod
    async def process_evidence(
        cls,
        claim_id: str,
        photos: List[Any],
        pdf_file: Optional[Any],
    ) -> Dict[str, Any]:
        cls.validate_evidence(photos, pdf_file)

        dirs = cls.ensure_claim_dirs(claim_id)
        errors: List[str] = []
        photos_metadata: List[Dict[str, Any]] = []

        for idx, photo in enumerate(photos, start=1):
            try:
                filename = cls._safe_filename(photo.filename, f"foto_{idx}")
                stored_name = f"foto_{idx:02d}{os.path.splitext(filename)[1].lower()}"
                target_path = os.path.join(dirs["images"], stored_name)
                size_bytes = cls._save_upload_file(photo, target_path, cls.MAX_PHOTO_BYTES)

                meta = cls._extract_image_metadata(target_path)
                meta.update({
                    "original_name": photo.filename,
                    "stored_name": stored_name,
                    "size_bytes": size_bytes,
                })
                photos_metadata.append(meta)
            except Exception as exc:
                errors.append(f"foto_{idx}: {str(exc)}")

        pdf_metadata = None
        if pdf_file is not None:
            try:
                pdf_name = cls._safe_filename(pdf_file.filename, "preforma.pdf")
                stored_pdf = f"preforma{os.path.splitext(pdf_name)[1].lower()}"
                pdf_path = os.path.join(dirs["docs"], stored_pdf)
                pdf_size = cls._save_upload_file(pdf_file, pdf_path, cls.MAX_PDF_BYTES)

                pdf_metadata = cls._extract_document_metadata(pdf_path)
                pdf_metadata.update({
                    "original_name": pdf_file.filename,
                    "stored_name": stored_pdf,
                    "size_bytes": pdf_size,
                })
            except Exception as exc:
                errors.append(f"pdf: {str(exc)}")

        summary = cls._build_summary(photos_metadata, pdf_metadata)

        metadata_payload = {
            "claim_id": claim_id,
            "created_at": datetime.utcnow().isoformat(),
            "photos": photos_metadata,
            "pdf": pdf_metadata,
            "summary": summary,
            "errors": errors,
        }

        metadata_path = os.path.join(dirs["base"], "metadata.json")
        with open(metadata_path, "w", encoding="utf-8") as handle:
            json.dump(metadata_payload, handle, indent=2, ensure_ascii=False)

        return summary

    @classmethod
    def load_evidence_summary(cls, claim_id: str) -> Optional[Dict[str, Any]]:
        metadata_path = os.path.join(cls.BASE_DIR, claim_id, "metadata.json")
        if not os.path.exists(metadata_path):
            return None
        try:
            with open(metadata_path, "r", encoding="utf-8") as handle:
                payload = json.load(handle)
            return payload.get("summary")
        except Exception:
            return None

    @classmethod
    def format_summary_text(cls, summary: Optional[Dict[str, Any]]) -> str:
        if not summary:
            return "No hay evidencias registradas para este siniestro."

        parts = [
            f"Fotos: {summary.get('photo_count', 0)}",
            f"GPS detectado: {summary.get('photos_with_gps', 0)}",
        ]

        models = summary.get("camera_models") or []
        if models:
            parts.append(f"Camaras: {', '.join(models)}")

        gps_points = summary.get("gps_points") or []
        if gps_points:
            sample = gps_points[:3]
            coord_text = "; ".join(
                f"{p.get('lat'):.6f},{p.get('lon'):.6f}"
                for p in sample
                if p.get("lat") is not None and p.get("lon") is not None
            )
            if coord_text:
                parts.append(f"Coordenadas: {coord_text}")

        doc_summary = summary.get("document") or summary.get("pdf") or {}
        if doc_summary:
            doc_type = doc_summary.get("doc_type") or "DOCUMENTO"
            parts.append(f"Documento: {doc_type} | paginas: {doc_summary.get('page_count', 0)}")
            if doc_type == "EXCEL":
                parts.append(f"Hojas: {doc_summary.get('sheet_count', 0)}")
            if doc_summary.get("has_text"):
                excerpt = doc_summary.get("text_excerpt", "")
                excerpt = re.sub(r"\s+", " ", excerpt).strip()
                if len(excerpt) > 220:
                    excerpt = excerpt[:220] + "..."
                if excerpt:
                    parts.append(f"Extracto: {excerpt}")

        return " | ".join(parts)

    @classmethod
    def _is_allowed_image(cls, upload: Any) -> bool:
        if upload.content_type in cls.ALLOWED_IMAGE_TYPES:
            return True
        ext = os.path.splitext(upload.filename or "")[1].lower()
        return ext in cls.ALLOWED_IMAGE_EXTS

    @classmethod
    def _is_allowed_pdf(cls, upload: Any) -> bool:
        if upload.content_type in cls.ALLOWED_PDF_TYPES:
            return True
        ext = os.path.splitext(upload.filename or "")[1].lower()
        return ext in cls.ALLOWED_PDF_EXTS

    @classmethod
    def _get_upload_size(cls, upload: Any) -> int:
        try:
            upload.file.seek(0, os.SEEK_END)
            size = upload.file.tell()
            upload.file.seek(0)
            return size
        except Exception:
            return 0

    @classmethod
    def _save_upload_file(cls, upload: Any, target_path: str, max_bytes: int) -> int:
        upload.file.seek(0)
        with open(target_path, "wb") as handle:
            shutil.copyfileobj(upload.file, handle)
        size = os.path.getsize(target_path)
        if size > max_bytes:
            os.remove(target_path)
            raise ValueError("Archivo excede el tamanio permitido.")
        return size

    @classmethod
    def _safe_filename(cls, filename: Optional[str], fallback: str) -> str:
        name = filename or fallback
        safe = re.sub(r"[^a-zA-Z0-9._-]", "_", name)
        return safe or fallback

    @classmethod
    def _extract_image_metadata(cls, path: str) -> Dict[str, Any]:
        meta = {
            "format": None,
            "width": None,
            "height": None,
            "exif_available": False,
        }

        if not PIL_AVAILABLE:
            meta["exif_available"] = False
            meta["exif_error"] = "PIL no disponible"
            return meta

        try:
            with Image.open(path) as img:
                meta["format"] = img.format
                meta["width"] = img.width
                meta["height"] = img.height

                exif_raw = img._getexif() or {}
                if not exif_raw:
                    return meta

                exif = {ExifTags.TAGS.get(k, k): v for k, v in exif_raw.items()}
                meta["exif_available"] = True
                meta["camera_make"] = exif.get("Make")
                meta["camera_model"] = exif.get("Model")
                meta["date_time_original"] = exif.get("DateTimeOriginal") or exif.get("DateTime")

                gps_info = exif.get("GPSInfo")
                if gps_info:
                    gps_parsed = cls._parse_gps_info(gps_info)
                    if gps_parsed:
                        meta["gps"] = gps_parsed
        except Exception as exc:
            meta["exif_error"] = str(exc)

        return meta

    @classmethod
    def _parse_gps_info(cls, gps_info: Any) -> Optional[Dict[str, Any]]:
        if not ExifTags:
            return None

        gps_data = {ExifTags.GPSTAGS.get(k, k): v for k, v in gps_info.items()}
        lat = cls._convert_gps(gps_data.get("GPSLatitude"), gps_data.get("GPSLatitudeRef"))
        lon = cls._convert_gps(gps_data.get("GPSLongitude"), gps_data.get("GPSLongitudeRef"))
        alt = cls._convert_altitude(gps_data.get("GPSAltitude"), gps_data.get("GPSAltitudeRef"))

        if lat is None or lon is None:
            return None

        return {"lat": lat, "lon": lon, "alt": alt}

    @classmethod
    def _convert_gps(cls, values: Any, ref: Optional[str]) -> Optional[float]:
        if not values:
            return None
        try:
            d = cls._rational_to_float(values[0])
            m = cls._rational_to_float(values[1])
            s = cls._rational_to_float(values[2])
            coord = d + (m / 60.0) + (s / 3600.0)
            if ref in ["S", "W"]:
                coord = -coord
            return coord
        except Exception:
            return None

    @classmethod
    def _convert_altitude(cls, value: Any, ref: Optional[int]) -> Optional[float]:
        if value is None:
            return None
        try:
            alt = cls._rational_to_float(value)
            if ref == 1:
                alt = -alt
            return alt
        except Exception:
            return None

    @classmethod
    def _rational_to_float(cls, value: Any) -> float:
        try:
            return float(value[0]) / float(value[1])
        except Exception:
            return float(value)

    @classmethod
    def extraer_fechas_del_texto(cls, text: str) -> List[str]:
        """
        Busca y extrae fechas en español en formato numérico y textual del texto.
        Retorna una lista de cadenas de fecha en formato ISO YYYY-MM-DD ordenadas.
        """
        import re
        from datetime import datetime
        
        dates = []
        
        # 1. Regex para formatos numéricos como DD/MM/YYYY o DD-MM-YYYY
        pattern_dmy = re.compile(r"\b(\d{1,2})[/-](\d{1,2})[/-](\d{4})\b")
        for match in pattern_dmy.finditer(text):
            d, m, y = match.groups()
            try:
                dt = datetime(int(y), int(m), int(d))
                dates.append(dt.date())
            except ValueError:
                try:
                    dt = datetime(int(y), int(d), int(m))
                    dates.append(dt.date())
                except ValueError:
                    pass

        # 2. Regex para formatos numéricos como YYYY-MM-DD o YYYY/MM/DD
        pattern_ymd = re.compile(r"\b(\d{4})[/-](\d{1,2})[/-](\d{1,2})\b")
        for match in pattern_ymd.finditer(text):
            y, m, d = match.groups()
            try:
                dt = datetime(int(y), int(m), int(d))
                dates.append(dt.date())
            except ValueError:
                pass

        # 3. Regex para fechas con nombre del mes en español
        meses = {
            "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
            "julio": 7, "agosto": 8, "septiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12
        }
        pattern_text = re.compile(
            r"\b(\d{1,2})\s+de\s+(enero|febrero|marzo|abril|mayo|junio|julio|agosto|septiembre|octubre|noviembre|diciembre)\s+de\s+(\d{4})\b",
            re.IGNORECASE
        )
        for match in pattern_text.finditer(text):
            d, mes_str, y = match.groups()
            m = meses.get(mes_str.lower())
            if m:
                try:
                    dt = datetime(int(y), m, int(d))
                    dates.append(dt.date())
                except ValueError:
                    pass

        # Eliminar duplicados manteniendo el orden
        seen = set()
        unique_dates = []
        for d in dates:
            if d not in seen:
                seen.add(d)
                unique_dates.append(d.isoformat())
        
        return sorted(unique_dates)

    @classmethod
    def extraer_montos_del_texto(cls, text: str) -> List[float]:
        """
        Busca y extrae montos de dinero del texto, especialmente aquellos cercanos a palabras como 'TOTAL' o 'SUBTOTAL'.
        Soporta enteros (7436) y decimales (10282.15) en líneas que contienen palabras clave de totales.
        También soporta el formato de celda separado por pipe (|) generado por el extractor de Excel.
        """
        import re
        montos = []
        lines = text.split("\n")
        for line in lines:
            line_upper = line.upper()
            # Detectar si la línea habla de un total/subtotal
            is_total_line = (
                "TOTAL" in line_upper
                or "SUBTOTAL" in line_upper
                or "SUB-TOTAL" in line_upper
                or "SUB TOTAL" in line_upper
            )
            if is_total_line:
                # Limpiar la línea: quitar signos de moneda, espacios internos múltiples
                clean_line = line.replace("$", "").replace(",", "")
                # Regex ampliado: captura tanto decimales (10282.15) como enteros (7436)
                # Al final de la línea o separados por pipes, puntos, espacios
                matches = re.findall(
                    r"\b(\d{1,3}(?:\.\d{3})*(?:\.\d+)?|\d+(?:\.\d+)?)\b",
                    clean_line
                )
                for match in matches:
                    try:
                        val = float(match.replace(",", ""))
                        # Solo aceptar valores >= 100 (descartar cantidades como 1, 15, 15%)
                        if val >= 100.0:
                            montos.append(val)
                    except ValueError:
                        pass
        return list(set(montos))


    @classmethod
    def _extract_pdf_metadata(cls, path: str) -> Dict[str, Any]:
        meta = {
            "doc_type": "PDF",
            "page_count": 0,
            "metadata": {},
            "has_text": False,
            "text_excerpt": "",
        }

        if not PDF_AVAILABLE:
            meta["pdf_error"] = "PyPDF2 no disponible"
            return meta

        try:
            reader = PdfReader(path)
            meta["page_count"] = len(reader.pages)
            if reader.metadata:
                meta["metadata"] = {str(k): str(v) for k, v in reader.metadata.items()}

            pages_to_read = min(2, len(reader.pages))
            text_chunks = []
            for idx in range(pages_to_read):
                try:
                    text = reader.pages[idx].extract_text() or ""
                except Exception:
                    text = ""
                if text:
                    text_chunks.append(text.strip())

            text_full = "\n".join(text_chunks).strip()
            meta["has_text"] = bool(text_full)
            meta["text_excerpt"] = text_full[:4000]
            
            # Extraer fechas y montos
            detected_dates = cls.extraer_fechas_del_texto(text_full)
            detected_montos = cls.extraer_montos_del_texto(text_full)
            meta["detected_dates"] = detected_dates
            meta["detected_montos"] = detected_montos
            
            # Estructuración uniforme de sheets para PDFs
            meta["sheets"] = [{
                "sheet_name": "PDF_DOCUMENT",
                "text_excerpt": text_full[:4000],
                "detected_dates": detected_dates,
                "detected_montos": detected_montos
            }]
        except Exception as exc:
            meta["pdf_error"] = str(exc)

        return meta

    @classmethod
    def _extract_word_metadata(cls, path: str) -> Dict[str, Any]:
        meta = {
            "doc_type": "WORD",
            "page_count": 0,
            "metadata": {},
            "has_text": False,
            "text_excerpt": "",
        }

        ext = os.path.splitext(path)[1].lower()
        if ext == ".doc" and not DOCX_AVAILABLE:
            meta["doc_error"] = "Extraccion .doc requiere herramientas externas"
            return meta

        if not DOCX_AVAILABLE:
            meta["doc_error"] = "python-docx no disponible"
            return meta

        if ext != ".docx":
            meta["doc_error"] = "Extraccion .doc no soportada con python-docx"
            return meta

        try:
            doc = DocxDocument(path)
            paragraphs = [p.text.strip() for p in doc.paragraphs if p.text and p.text.strip()]
            text_full = "\n".join(paragraphs).strip()
            meta["has_text"] = bool(text_full)
            meta["text_excerpt"] = text_full[:4000]

            # Extraer fechas y montos
            detected_dates = cls.extraer_fechas_del_texto(text_full)
            detected_montos = cls.extraer_montos_del_texto(text_full)
            meta["detected_dates"] = detected_dates
            meta["detected_montos"] = detected_montos

            # Estructuración uniforme de sheets para Word
            meta["sheets"] = [{
                "sheet_name": "WORD_DOCUMENT",
                "text_excerpt": text_full[:4000],
                "detected_dates": detected_dates,
                "detected_montos": detected_montos
            }]

            props = doc.core_properties
            meta["metadata"] = {
                "title": props.title,
                "author": props.author,
                "created": str(props.created) if props.created else None,
                "modified": str(props.modified) if props.modified else None,
            }
        except Exception as exc:
            meta["doc_error"] = str(exc)

        return meta

    @classmethod
    def _extract_excel_metadata(cls, path: str) -> Dict[str, Any]:
        meta = {
            "doc_type": "EXCEL",
            "page_count": 0,
            "sheet_count": 0,
            "metadata": {},
            "has_text": False,
            "text_excerpt": "",
        }

        ext = os.path.splitext(path)[1].lower()
        if ext == ".xlsx":
            if not OPENPYXL_AVAILABLE:
                meta["doc_error"] = "openpyxl no disponible"
                return meta
            try:
                workbook = load_workbook(path, read_only=True, data_only=True)
                meta["sheet_count"] = len(workbook.sheetnames)
                meta["metadata"] = {"sheet_names": workbook.sheetnames}

                all_sheets_text = []
                sheets_info = []
                for sheet_name in workbook.sheetnames:
                    ws = workbook[sheet_name]
                    lines = []
                    raw_numeric_montos = []
                    # Extracción profunda: hasta 100 filas y 20 columnas (para no perder totales en col 13+)
                    for row in ws.iter_rows(min_row=1, max_row=100, max_col=20, values_only=True):
                        cells = ["" if v is None else str(v).strip() for v in row]
                        line = " | ".join([c for c in cells if c != ""])
                        if line:
                            lines.append(line)
                        
                        # Extraer montos numéricos crudos de celdas que parecen montos y están en líneas con palabras clave de totales
                        row_upper_strs = ["" if v is None else str(v).upper() for v in row]
                        row_has_total_keyword = any(
                            keyword in s
                            for s in row_upper_strs
                            for keyword in ["TOTAL", "SUBTOTAL", "SUB-TOTAL", "SUB TOTAL"]
                        )
                        if row_has_total_keyword:
                            for val in row:
                                if isinstance(val, (int, float)) and val >= 100.0:
                                    raw_numeric_montos.append(float(val))
                                    
                    if lines:
                        sheet_text = "\n".join(lines)
                        detected_dates = cls.extraer_fechas_del_texto(sheet_text)
                        text_montos = cls.extraer_montos_del_texto(sheet_text)
                        detected_montos = sorted(list(set(text_montos + raw_numeric_montos)))
                        sheets_info.append({
                            "sheet_name": sheet_name,
                            "text_excerpt": sheet_text[:3000],
                            "detected_dates": detected_dates,
                            "detected_montos": detected_montos
                        })
                        all_sheets_text.append(f"--- HOJA: {sheet_name} ---\n" + sheet_text)

                text_full = "\n\n".join(all_sheets_text).strip()
                meta["has_text"] = bool(text_full)
                meta["text_excerpt"] = text_full[:4000]
                meta["sheets"] = sheets_info
                
                # Agregar fechas y montos agregados a nivel global de metadata
                global_dates = []
                global_montos = []
                for s in sheets_info:
                    global_dates.extend(s["detected_dates"])
                    global_montos.extend(s["detected_montos"])
                meta["detected_dates"] = sorted(list(set(global_dates)))
                meta["detected_montos"] = sorted(list(set(global_montos)))
            except Exception as exc:
                meta["doc_error"] = str(exc)
            return meta

        if ext == ".xls":
            if not XLRD_AVAILABLE:
                meta["doc_error"] = "xlrd no disponible"
                return meta
            try:
                workbook = xlrd.open_workbook(path)
                meta["sheet_count"] = workbook.nsheets
                meta["metadata"] = {"sheet_names": workbook.sheet_names()}

                all_sheets_text = []
                sheets_info = []
                for sheet_idx in range(workbook.nsheets):
                    sheet = workbook.sheet_by_index(sheet_idx)
                    sheet_name = workbook.sheet_names()[sheet_idx]
                    lines = []
                    raw_numeric_montos = []
                    max_row = min(sheet.nrows, 100)
                    max_col = min(sheet.ncols, 20)
                    for r in range(max_row):
                        cells = [sheet.cell_value(r, c) for c in range(max_col)]
                        cells_str = ["" if v is None else str(v).strip() for v in cells]
                        line = " | ".join([c for c in cells_str if c != ""])
                        if line:
                            lines.append(line)
                        
                        # Extraer montos numéricos crudos de celdas
                        row_upper_strs = ["" if v is None else str(v).upper() for v in cells]
                        row_has_total_keyword = any(
                            keyword in s
                            for s in row_upper_strs
                            for keyword in ["TOTAL", "SUBTOTAL", "SUB-TOTAL", "SUB TOTAL"]
                        )
                        if row_has_total_keyword:
                            for val in cells:
                                if isinstance(val, (int, float)) and val >= 100.0:
                                    raw_numeric_montos.append(float(val))
                                    
                    if lines:
                        sheet_text = "\n".join(lines)
                        detected_dates = cls.extraer_fechas_del_texto(sheet_text)
                        text_montos = cls.extraer_montos_del_texto(sheet_text)
                        detected_montos = sorted(list(set(text_montos + raw_numeric_montos)))
                        sheets_info.append({
                            "sheet_name": sheet_name,
                            "text_excerpt": sheet_text[:3000],
                            "detected_dates": detected_dates,
                            "detected_montos": detected_montos
                        })
                        all_sheets_text.append(f"--- HOJA: {sheet_name} ---\n" + sheet_text)

                text_full = "\n\n".join(all_sheets_text).strip()
                meta["has_text"] = bool(text_full)
                meta["text_excerpt"] = text_full[:4000]
                meta["sheets"] = sheets_info
                
                # Agregar fechas y montos agregados a nivel global de metadata
                global_dates = []
                global_montos = []
                for s in sheets_info:
                    global_dates.extend(s["detected_dates"])
                    global_montos.extend(s["detected_montos"])
                meta["detected_dates"] = sorted(list(set(global_dates)))
                meta["detected_montos"] = sorted(list(set(global_montos)))
            except Exception as exc:
                meta["doc_error"] = str(exc)
            return meta

        meta["doc_error"] = "Formato de Excel no soportado"
        return meta

    @classmethod
    def _extract_document_metadata(cls, path: str) -> Dict[str, Any]:
        doc_type = cls._get_document_type(path)
        if doc_type == "PDF":
            return cls._extract_pdf_metadata(path)
        if doc_type == "WORD":
            return cls._extract_word_metadata(path)
        if doc_type == "EXCEL":
            return cls._extract_excel_metadata(path)
        return {
            "doc_type": doc_type,
            "page_count": 0,
            "sheet_count": 0,
            "metadata": {},
            "has_text": False,
            "text_excerpt": "",
        }

    @classmethod
    def _get_document_type(cls, path: str) -> str:
        ext = os.path.splitext(path)[1].lower()
        if ext == ".pdf":
            return "PDF"
        if ext in {".doc", ".docx"}:
            return "WORD"
        if ext in {".xls", ".xlsx"}:
            return "EXCEL"
        return "DOCUMENTO"

    @classmethod
    def _build_summary(
        cls,
        photos_metadata: List[Dict[str, Any]],
        pdf_metadata: Optional[Dict[str, Any]],
    ) -> Dict[str, Any]:
        gps_points = []
        camera_models = set()
        capture_dates = set()

        for item in photos_metadata:
            gps = item.get("gps")
            if gps:
                gps_points.append({
                    "lat": gps.get("lat"),
                    "lon": gps.get("lon"),
                    "alt": gps.get("alt"),
                    "source": item.get("stored_name"),
                })
            if item.get("camera_model"):
                camera_models.add(str(item.get("camera_model")))
            if item.get("date_time_original"):
                capture_dates.add(str(item.get("date_time_original")))

        summary_doc = {}
        if pdf_metadata:
            summary_doc = {
                "doc_type": pdf_metadata.get("doc_type", "DOCUMENTO"),
                "page_count": pdf_metadata.get("page_count", 0),
                "sheet_count": pdf_metadata.get("sheet_count", 0),
                "has_text": pdf_metadata.get("has_text", False),
                "text_excerpt": pdf_metadata.get("text_excerpt", ""),
                "sheets": pdf_metadata.get("sheets", []),
                "detected_dates": pdf_metadata.get("detected_dates", []),
                "detected_montos": pdf_metadata.get("detected_montos", []),
            }

        return {
            "photo_count": len(photos_metadata),
            "photos_with_gps": len(gps_points),
            "gps_points": gps_points,
            "camera_models": sorted(camera_models),
            "capture_dates": sorted(capture_dates),
            "document": summary_doc,
        }
